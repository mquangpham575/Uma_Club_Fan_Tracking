import sys
import os
import openpyxl

def inspect_excel_structure():
    path = r"c:\Project\uma_tracker\uma_tracking\FEBRUARY DATA UMA - ENDLESS.xlsx"
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
        
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['TOP PERFORMER + CLUB RANK']
    
    print("Merges:")
    for r in sorted(list(ws.merged_cells.ranges), key=lambda x: (x.min_row, x.min_col)):
        print("Merged range:", r)
        
    # Check specific cells in row 1
    print("\n--- Row 1 Cells ---")
    for col in range(1, 15):
        cell = ws.cell(row=1, column=col)
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}1) value: '{cell.value}', fill: '{cell.fill.fgColor.rgb if cell.fill else None}'")
        
    # Check specific cells in row 2
    print("\n--- Row 2 Cells ---")
    for col in range(1, 15):
        cell = ws.cell(row=2, column=col)
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}2) value: '{cell.value}', fill: '{cell.fill.fgColor.rgb if cell.fill else None}'")

if __name__ == '__main__':
    inspect_excel_structure()
